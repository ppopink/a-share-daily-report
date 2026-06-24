"""
选股报告生成器

生成面向日常查看的 Excel 和 HTML 报告；CSV 继续作为机器可读结果。
"""

import datetime
import html
import json
import os
import re
import shutil
import subprocess
import tempfile
from typing import Optional

import pandas as pd

import config as cfg
from utils import output_day_dir


DISPLAY_COLUMNS = [
    ("rank", "排名"),
    ("code", "代码"),
    ("name", "名称"),
    ("industry", "行业"),
    ("price", "现价"),
    ("pct_change", "涨跌幅%"),
    ("total_score", "总分"),
    ("trend_score", "趋势"),
    ("above_ma20_days_20", "20日线上天数"),
    ("ma20_slope_10_pct", "MA20十日斜率%"),
    ("close_ma20_ratio", "股价/MA20"),
    ("volume_score", "量能"),
    ("dmi_score", "DMI"),
    ("macd_score", "MACD"),
    ("expma_score", "EXPMA"),
    ("money_flow_score", "资金"),
    ("sector_score", "板块"),
    ("leading_score", "领涨"),
    ("adx", "ADX"),
    ("plus_di", "+DI"),
    ("minus_di", "-DI"),
    ("vol_ratio", "量比"),
    ("main_net_ratio", "主力净流入比"),
    ("sector_rank_pct", "板块排名分位"),
    ("stock_rank_in_sector", "板块内排名分位"),
    ("entry_label", "入场时机"),
    ("planned_holding_days", "计划持有日"),
    ("stop_loss_price", "止损价"),
    ("take_profit_1_price", "第一止盈"),
    ("take_profit_2_price", "第二止盈"),
    ("trailing_stop_price", "移动止盈"),
    ("risk_reward_ratio", "盈亏比"),
    ("exit_signal", "退出计划"),
    ("day1_stop_loss_price", "第1日止损"),
    ("day1_take_profit_price", "第1日止盈"),
    ("day1_exit_plan", "第1日计划"),
    ("day2_stop_loss_price", "第2日止损"),
    ("day2_take_profit_price", "第2日止盈"),
    ("day2_exit_plan", "第2日计划"),
    ("day3_stop_loss_price", "第3日止损"),
    ("day3_take_profit_price", "第3日止盈"),
    ("day3_exit_plan", "第3日计划"),
]

SCORE_PARTS = [
    ("trend_score", "趋势"),
    ("volume_score", "量能"),
    ("dmi_score", "DMI"),
    ("macd_score", "MACD"),
    ("expma_score", "EXPMA"),
    ("money_flow_score", "资金"),
    ("sector_score", "板块"),
    ("leading_score", "领涨"),
]


def _date_key(trade_date=None) -> str:
    if trade_date is None:
        return datetime.date.today().strftime("%Y%m%d")
    if isinstance(trade_date, str):
        return pd.to_datetime(trade_date).strftime("%Y%m%d")
    return trade_date.strftime("%Y%m%d")


def _date_label(trade_date=None) -> str:
    if trade_date is None:
        return datetime.date.today().strftime("%Y-%m-%d")
    if isinstance(trade_date, str):
        return pd.to_datetime(trade_date).strftime("%Y-%m-%d")
    return trade_date.strftime("%Y-%m-%d")


def _safe_float(value, default=0.0) -> float:
    try:
        if pd.isna(value):
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def _fmt(value, digits=2):
    if pd.isna(value):
        return ""
    if isinstance(value, float):
        return f"{value:.{digits}f}"
    return value


def _load_csv_if_exists(path: str) -> pd.DataFrame:
    if path and os.path.exists(path):
        try:
            if os.path.getsize(path) <= 4:
                return pd.DataFrame()
            return pd.read_csv(path)
        except pd.errors.EmptyDataError:
            return pd.DataFrame()
    return pd.DataFrame()


def _dated_or_legacy_path(filename: str, date_key: str) -> str:
    dated = os.path.join(cfg.OUTPUT_DIR, date_key, filename)
    if os.path.exists(dated):
        return dated
    return os.path.join(cfg.OUTPUT_DIR, filename)


def _prepare_pick_df(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if "rank" not in out.columns:
        out.insert(0, "rank", range(1, len(out) + 1))
    for col, _ in DISPLAY_COLUMNS:
        if col not in out.columns:
            out[col] = ""
    return out[[col for col, _ in DISPLAY_COLUMNS]]


def _summary_stats(df: pd.DataFrame) -> dict:
    if df.empty:
        return {
            "pick_count": 0,
            "top_name": "",
            "top_score": 0,
            "avg_score": 0,
            "avg_adx": 0,
            "avg_vol_ratio": 0,
            "risk_count": 0,
        }
    return {
        "pick_count": int(len(df)),
        "top_name": f"{df.iloc[0].get('name', '')}({df.iloc[0].get('code', '')})",
        "top_score": _safe_float(df.iloc[0].get("total_score", 0)),
        "avg_score": _safe_float(pd.to_numeric(df.get("total_score"), errors="coerce").mean()),
        "avg_adx": _safe_float(pd.to_numeric(df.get("adx"), errors="coerce").mean()),
        "avg_vol_ratio": _safe_float(pd.to_numeric(df.get("vol_ratio"), errors="coerce").mean()),
        "risk_count": int(df.get("entry_label", pd.Series([], dtype=str)).astype(str).str.contains("追高|远离|风险|连涨", na=False).sum()),
    }


def _tech_diag_path(date_key: str) -> str:
    return _dated_or_legacy_path(f"tech_diagnostic_{date_key}.csv", date_key)


def _filter_log_path(date_key: str) -> str:
    return _dated_or_legacy_path(f"filter_log_{date_key}.csv", date_key)


def _accuracy_summary_path(date_key: str) -> str:
    return _dated_or_legacy_path(f"prediction_accuracy_summary_{date_key}.csv", date_key)


def _accuracy_ic_path(date_key: str) -> str:
    return _dated_or_legacy_path(f"prediction_feature_ic_{date_key}.csv", date_key)


def generate_excel_report(
    pick_df: pd.DataFrame,
    trade_date=None,
    output_dir: str = None,
) -> Optional[str]:
    try:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, PieChart, Reference, ScatterChart, Series
        from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, CellIsRule
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError:
        print("[报告] 未安装 openpyxl，跳过 Excel 报告。可运行: pip install openpyxl")
        return None

    output_dir = output_dir or output_day_dir(trade_date)
    os.makedirs(output_dir, exist_ok=True)
    date_key = _date_key(trade_date)
    date_label = _date_label(trade_date)
    xlsx_path = os.path.join(output_dir, f"stock_pick_{date_key}.xlsx")

    picks = _prepare_pick_df(pick_df)
    tech_diag = _load_csv_if_exists(_tech_diag_path(date_key))
    filter_log = _load_csv_if_exists(_filter_log_path(date_key))
    accuracy = _load_csv_if_exists(_accuracy_summary_path(date_key))
    feature_ic = _load_csv_if_exists(_accuracy_ic_path(date_key))
    stats = _summary_stats(picks)

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "摘要"
    ws_picks = wb.create_sheet("选股详情")
    ws_diag = wb.create_sheet("技术诊断")
    ws_filter = wb.create_sheet("筛选漏斗")
    ws_accuracy = wb.create_sheet("准确性评估")

    dark = "1F4E78"
    blue = "D9EAF7"
    green = "E2F0D9"
    red = "FCE4D6"
    gray = "F2F2F2"
    border = Border(bottom=Side(style="thin", color="D9E2F3"))

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

    # 摘要
    ws_summary.merge_cells("A1:H1")
    ws_summary["A1"] = f"K线选股日报 - {date_label}"
    ws_summary["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    ws_summary["A1"].fill = PatternFill("solid", fgColor=dark)
    ws_summary["A1"].alignment = Alignment(horizontal="center")
    summary_cards = [
        ("入选数量", stats["pick_count"], "只"),
        ("最高分股票", stats["top_name"], ""),
        ("最高分", stats["top_score"], ""),
        ("平均分", stats["avg_score"], ""),
        ("平均ADX", stats["avg_adx"], ""),
        ("平均量比", stats["avg_vol_ratio"], ""),
        ("风险提示数", stats["risk_count"], "只"),
    ]
    for i, (label, value, unit) in enumerate(summary_cards, start=3):
        ws_summary[f"A{i}"] = label
        ws_summary[f"B{i}"] = value
        ws_summary[f"C{i}"] = unit
        ws_summary[f"A{i}:C{i}"][0][0].fill = PatternFill("solid", fgColor=blue)
        for cell in ws_summary[f"A{i}:C{i}"][0]:
            cell.border = border
            cell.alignment = Alignment(vertical="center")
        ws_summary[f"A{i}"].font = Font(bold=True)

    note_row = 12
    ws_summary[f"A{note_row}"] = "阅读提示"
    ws_summary[f"A{note_row}"].font = Font(bold=True, color="1F4E78")
    ws_summary[f"A{note_row + 1}"] = "1. 本报告仅用于规则选股跟踪，不构成投资建议。"
    ws_summary[f"A{note_row + 2}"] = "2. 总分越高代表规则综合排序越靠前；仍需结合流动性、涨跌停、停牌和市场环境。"
    ws_summary[f"A{note_row + 3}"] = "3. 准确性评估需要未来交易日样本成熟后才会出现有效统计。"

    # 选股详情
    headers = [label for _, label in DISPLAY_COLUMNS]
    ws_picks.append(headers)
    for _, row in picks.iterrows():
        ws_picks.append([row.get(col, "") for col, _ in DISPLAY_COLUMNS])
    for cell in ws_picks[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.alignment = Alignment(horizontal="center")
    ws_picks.freeze_panes = "A2"
    ws_picks.auto_filter.ref = ws_picks.dimensions
    table_ref = f"A1:{get_column_letter(ws_picks.max_column)}{max(ws_picks.max_row, 2)}"
    tab = Table(displayName="StockPicks", ref=table_ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws_picks.add_table(tab)

    col_index = {label: idx + 1 for idx, label in enumerate(headers)}
    if ws_picks.max_row > 1:
        score_col = get_column_letter(col_index["总分"])
        pct_col = get_column_letter(col_index["涨跌幅%"])
        adx_col = get_column_letter(col_index["ADX"])
        vol_col = get_column_letter(col_index["量比"])
        ws_picks.conditional_formatting.add(
            f"{score_col}2:{score_col}{ws_picks.max_row}",
            ColorScaleRule(start_type="min", start_color="F8696B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="63BE7B"),
        )
        ws_picks.conditional_formatting.add(
            f"{pct_col}2:{pct_col}{ws_picks.max_row}",
            CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor=green)),
        )
        ws_picks.conditional_formatting.add(
            f"{pct_col}2:{pct_col}{ws_picks.max_row}",
            CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor=red)),
        )
        ws_picks.conditional_formatting.add(
            f"{adx_col}2:{adx_col}{ws_picks.max_row}",
            DataBarRule(start_type="num", start_value=0, end_type="num", end_value=80, color="5B9BD5"),
        )
        ws_picks.conditional_formatting.add(
            f"{vol_col}2:{vol_col}{ws_picks.max_row}",
            DataBarRule(start_type="num", start_value=0, end_type="num", end_value=3, color="70AD47"),
        )

    # 技术诊断
    _write_dataframe(ws_diag, tech_diag, title="技术条件诊断", header_color=dark)
    # 筛选漏斗
    _write_dataframe(ws_filter, filter_log, title="筛选漏斗", header_color=dark)
    # 准确性评估
    if accuracy.empty:
        ws_accuracy["A1"] = "准确性评估"
        ws_accuracy["A2"] = "暂无成熟样本；请在未来交易日后运行 python main.py --evaluate --eval-fetch-missing。"
    else:
        _write_dataframe(ws_accuracy, accuracy, title="准确性评估汇总", header_color=dark)
        start_row = ws_accuracy.max_row + 3
        if not feature_ic.empty:
            ws_accuracy[f"A{start_row}"] = "指标预测力 IC"
            ws_accuracy[f"A{start_row}"].font = Font(bold=True, color="1F4E78")
            _append_dataframe(ws_accuracy, feature_ic, start_row + 1, 1, header_color=dark)

    # 图表
    if ws_picks.max_row > 1:
        chart = BarChart()
        chart.title = "Top股票总分"
        chart.y_axis.title = "总分"
        chart.x_axis.title = "股票"
        data = Reference(ws_picks, min_col=col_index["总分"], min_row=1, max_row=ws_picks.max_row)
        cats = Reference(ws_picks, min_col=col_index["名称"], min_row=2, max_row=ws_picks.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 18
        ws_summary.add_chart(chart, "E3")

        scatter = ScatterChart()
        scatter.title = "ADX 与量比"
        scatter.x_axis.title = "量比"
        scatter.y_axis.title = "ADX"
        xvalues = Reference(ws_picks, min_col=col_index["量比"], min_row=2, max_row=ws_picks.max_row)
        yvalues = Reference(ws_picks, min_col=col_index["ADX"], min_row=2, max_row=ws_picks.max_row)
        series = Series(yvalues, xvalues, title="候选股")
        scatter.series.append(series)
        scatter.height = 8
        scatter.width = 18
        ws_summary.add_chart(scatter, "E20")

    industry_counts = picks["industry"].fillna("").astype(str)
    industry_counts = industry_counts[industry_counts != ""].value_counts().head(8)
    if not industry_counts.empty:
        start = 22
        ws_summary[f"A{start}"] = "行业分布"
        ws_summary[f"A{start}"].font = Font(bold=True, color="1F4E78")
        for i, (industry, count) in enumerate(industry_counts.items(), start=start + 1):
            ws_summary[f"A{i}"] = industry
            ws_summary[f"B{i}"] = int(count)
        pie = PieChart()
        pie.title = "行业分布"
        data = Reference(ws_summary, min_col=2, min_row=start + 1, max_row=start + len(industry_counts))
        cats = Reference(ws_summary, min_col=1, min_row=start + 1, max_row=start + len(industry_counts))
        pie.add_data(data)
        pie.set_categories(cats)
        pie.height = 7
        pie.width = 9
        ws_summary.add_chart(pie, "A32")

    for ws in wb.worksheets:
        _format_sheet(ws)

    wb.save(xlsx_path)
    print(f"[报告] Excel报告: {xlsx_path}")
    return xlsx_path


def _write_dataframe(ws, df: pd.DataFrame, title: str, header_color: str):
    from openpyxl.styles import Font

    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    if df.empty:
        ws["A2"] = "暂无数据"
        return
    _append_dataframe(ws, df, 3, 1, header_color)


def _append_dataframe(ws, df: pd.DataFrame, start_row: int, start_col: int, header_color: str):
    from openpyxl.styles import Alignment, Font, PatternFill

    for j, col in enumerate(df.columns, start=start_col):
        cell = ws.cell(start_row, j, col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=header_color)
        cell.alignment = Alignment(horizontal="center")
    for i, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        for j, col in enumerate(df.columns, start=start_col):
            val = row[col]
            if pd.isna(val):
                val = ""
            ws.cell(i, j, val)
    ws.freeze_panes = ws.cell(start_row + 1, start_col).coordinate
    ws.auto_filter.ref = f"{ws.cell(start_row, start_col).coordinate}:{ws.cell(start_row + len(df), start_col + len(df.columns) - 1).coordinate}"


def _format_sheet(ws):
    from openpyxl.styles import Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin", color="E7E6E6")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if cell.value not in (None, ""):
                cell.border = Border(bottom=thin)
    for col_cells in ws.columns:
        values = [str(c.value) for c in col_cells if c.value is not None]
        if not values:
            continue
        width = min(max(max(len(v) for v in values) + 2, 8), 28)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = width
    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 22


def _bar_svg(items, value_key="value", label_key="label", width=720, bar_height=26):
    items = list(items)
    if not items:
        return "<div class='empty'>暂无数据</div>"
    max_val = max(abs(_safe_float(x.get(value_key, 0))) for x in items) or 1
    rows = []
    for item in items:
        label = html.escape(str(item.get(label_key, "")))
        val = _safe_float(item.get(value_key, 0))
        pct = max(min(abs(val) / max_val, 1), 0)
        color = "#2E7D32" if val >= 0 else "#C62828"
        rows.append(
            f"<div class='bar-row'><span class='bar-label'>{label}</span>"
            f"<span class='bar-track'><span class='bar-fill' style='width:{pct*100:.1f}%;background:{color}'></span></span>"
            f"<span class='bar-value'>{val:.2f}</span></div>"
        )
    return "\n".join(rows)


def _score_stack_html(row: pd.Series) -> str:
    total = sum(max(_safe_float(row.get(col)), 0) for col, _ in SCORE_PARTS) or 1
    parts = []
    colors = ["#2F80ED", "#27AE60", "#9B51E0", "#F2994A", "#56CCF2", "#EB5757", "#6FCF97", "#BDBDBD"]
    for i, (col, label) in enumerate(SCORE_PARTS):
        val = max(_safe_float(row.get(col)), 0)
        if val <= 0:
            continue
        parts.append(
            f"<span title='{html.escape(label)} {val:.1f}' style='width:{val/total*100:.1f}%;background:{colors[i % len(colors)]}'></span>"
        )
    return "<div class='stack'>" + "".join(parts) + "</div>"


def _scatter_html(df: pd.DataFrame) -> str:
    if df.empty or "adx" not in df.columns or "vol_ratio" not in df.columns:
        return "<div class='empty'>暂无散点数据</div>"
    pts = []
    w, h = 520, 320
    xs = pd.to_numeric(df["vol_ratio"], errors="coerce").fillna(0)
    ys = pd.to_numeric(df["adx"], errors="coerce").fillna(0)
    max_x = max(xs.max(), 1)
    max_y = max(ys.max(), 1)
    for _, row in df.iterrows():
        x = 45 + _safe_float(row.get("vol_ratio")) / max_x * (w - 80)
        y = h - 35 - _safe_float(row.get("adx")) / max_y * (h - 70)
        name = html.escape(str(row.get("name", "")))
        score = _safe_float(row.get("total_score"))
        radius = 5 + min(max(score - 50, 0), 50) / 20
        pts.append(f"<circle cx='{x:.1f}' cy='{y:.1f}' r='{radius:.1f}'><title>{name} ADX {_safe_float(row.get('adx')):.1f} 量比 {_safe_float(row.get('vol_ratio')):.2f}</title></circle>")
    return (
        f"<svg class='scatter' viewBox='0 0 {w} {h}' role='img'>"
        f"<line x1='45' y1='{h-35}' x2='{w-25}' y2='{h-35}'/>"
        f"<line x1='45' y1='20' x2='45' y2='{h-35}'/>"
        f"<text x='{w/2}' y='{h-8}'>量比</text><text x='8' y='20'>ADX</text>"
        + "".join(pts)
        + "</svg>"
    )


def _build_share_sections(picks: pd.DataFrame, tech_diag: pd.DataFrame, stats: dict, date_label: str) -> tuple:
    top_rows = picks.head(5)
    top_names = [
        f"{row.get('name', '')}({row.get('code', '')})"
        for _, row in top_rows.iterrows()
    ]
    high_score_count = int((pd.to_numeric(picks.get("total_score"), errors="coerce") >= 80).sum()) if not picks.empty else 0
    risk_rows = picks[picks.get("entry_label", pd.Series([], dtype=str)).astype(str).str.contains("追高|远离|风险|连涨", na=False)].head(5)
    risk_names = [
        f"{row.get('name', '')}({row.get('entry_label', '')})"
        for _, row in risk_rows.iterrows()
    ]

    killer_text = "暂无技术诊断"
    if not tech_diag.empty and {"condition_name", "pass_rate"}.issubset(tech_diag.columns):
        diag = tech_diag.copy()
        diag = diag[~diag["condition_name"].astype(str).str.contains("最终", na=False)]
        diag["pass_rate_num"] = pd.to_numeric(diag["pass_rate"], errors="coerce")
        diag = diag.dropna(subset=["pass_rate_num"])
        if not diag.empty:
            row = diag.sort_values("pass_rate_num").iloc[0]
            killer_text = f"{row['condition_name']}，通过率 {row['pass_rate_num']:.1%}"

    conclusion_items = [
        f"今日入选 {stats['pick_count']} 只，最高分为 {stats['top_score']:.1f}。",
        f"Top5：{'、'.join(top_names) if top_names else '暂无'}。",
        f"80分以上股票 {high_score_count} 只，平均分 {stats['avg_score']:.1f}。",
        f"技术筛选最严格条件：{killer_text}。",
        f"风险提示股票 {stats['risk_count']} 只，需重点看入场时机。",
    ]
    conclusion_html = "<ul class='conclusion-list'>" + "".join(
        f"<li>{html.escape(item)}</li>" for item in conclusion_items
    ) + "</ul>"

    risk_html = "<div class='empty'>暂无明显追高/远离均线风险提示</div>"
    if risk_names:
        risk_html = "<ul class='risk-list'>" + "".join(
            f"<li>{html.escape(item)}</li>" for item in risk_names
        ) + "</ul>"

    share_text = "\n".join([
        f"{date_label} K线规则选股观察",
        f"入选：{stats['pick_count']}只；最高分：{stats['top_name']} {stats['top_score']:.1f}分",
        f"Top5：{'、'.join(top_names) if top_names else '暂无'}",
        f"技术最严格条件：{killer_text}",
        "仅供规则跟踪和学习交流，不构成投资建议；历史表现不代表未来收益。",
    ])
    return conclusion_html, risk_html, share_text


def generate_html_report(
    pick_df: pd.DataFrame,
    trade_date=None,
    output_dir: str = None,
) -> str:
    output_dir = output_dir or output_day_dir(trade_date)
    os.makedirs(output_dir, exist_ok=True)
    date_key = _date_key(trade_date)
    date_label = _date_label(trade_date)
    html_path = os.path.join(output_dir, f"stock_pick_report_{date_key}.html")

    picks = _prepare_pick_df(pick_df)
    stats = _summary_stats(picks)
    tech_diag = _load_csv_if_exists(_tech_diag_path(date_key))
    accuracy = _load_csv_if_exists(_accuracy_summary_path(date_key))

    top_items = [
        {"label": f"{row['rank']}. {row['name']}", "value": _safe_float(row["total_score"])}
        for _, row in picks.head(20).iterrows()
    ]
    funnel_items = []
    if not tech_diag.empty and {"condition_name", "pass_count"}.issubset(tech_diag.columns):
        funnel_items = [
            {"label": row["condition_name"], "value": _safe_float(row["pass_count"])}
            for _, row in tech_diag.iterrows()
        ]

    table_rows = []
    for _, row in picks.iterrows():
        pct = _safe_float(row.get("pct_change"))
        pct_cls = "up" if pct >= 0 else "down"
        table_rows.append(
            "<tr>"
            f"<td>{int(_safe_float(row.get('rank')))}</td>"
            f"<td>{html.escape(str(row.get('code', '')))}</td>"
            f"<td>{html.escape(str(row.get('name', '')))}</td>"
            f"<td>{_fmt(row.get('price'), 2)}</td>"
            f"<td>{html.escape(str(row.get('industry', '')))}</td>"
            f"<td>{_fmt(row.get('total_score'))}</td>"
            f"<td>{int(_safe_float(row.get('above_ma20_days_20')))}</td>"
            f"<td>{_fmt(row.get('ma20_slope_10_pct'), 2)}%</td>"
            f"<td class='{pct_cls}'>{pct:+.2f}%</td>"
            f"<td>{_fmt(row.get('adx'), 1)}</td>"
            f"<td>{_fmt(row.get('vol_ratio'), 2)}</td>"
            f"<td>{_score_stack_html(row)}</td>"
            f"<td>{html.escape(str(row.get('entry_label', '')))}</td>"
            f"<td>{html.escape(str(row.get('exit_signal', '')))}</td>"
            f"<td>{html.escape(str(row.get('day1_exit_plan', '')))}</td>"
            f"<td>{html.escape(str(row.get('day2_exit_plan', '')))}</td>"
            f"<td>{html.escape(str(row.get('day3_exit_plan', '')))}</td>"
            "</tr>"
        )

    accuracy_html = "<div class='empty'>暂无成熟准确性样本</div>"
    if not accuracy.empty:
        rows = []
        for _, row in accuracy.iterrows():
            rows.append(
                "<tr>"
                f"<td>{int(_safe_float(row.get('hold_days')))}</td>"
                f"<td>{int(_safe_float(row.get('sample_count')))}</td>"
                f"<td>{_safe_float(row.get('win_rate')):.1%}</td>"
                f"<td>{_safe_float(row.get('average_return_pct')):.2f}%</td>"
                f"<td>{_safe_float(row.get('average_excess_return_pct')):.2f}%</td>"
                "</tr>"
            )
        accuracy_html = "<table><thead><tr><th>持有日</th><th>样本</th><th>胜率</th><th>平均收益</th><th>平均超额</th></tr></thead><tbody>" + "".join(rows) + "</tbody></table>"

    conclusion_html, risk_html, share_text = _build_share_sections(picks, tech_diag, stats, date_label)
    generated_at = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    doc = f"""<!doctype html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<title>K线选股日报 {date_label}</title>
<style>
:root {{ --blue:#1f4e78; --light:#f5f8fb; --green:#16794c; --red:#b42318; --line:#e5e7eb; }}
body {{ margin:0; background:#f4f6f8; color:#172033; font-family:-apple-system,BlinkMacSystemFont,"Segoe UI","PingFang SC","Microsoft YaHei",Arial,sans-serif; }}
.wrap {{ max-width:1180px; margin:0 auto; padding:28px; }}
.hero {{ background:linear-gradient(135deg,#173b5f,#286da8); color:white; padding:28px 32px; border-radius:10px; }}
.hero h1 {{ margin:0 0 8px; font-size:28px; }}
.hero p {{ margin:0; opacity:.9; }}
.disclaimer {{ margin-top:16px; background:#fff7ed; color:#9a3412; border:1px solid #fed7aa; border-radius:8px; padding:12px 14px; font-weight:700; }}
.meta {{ display:flex; flex-wrap:wrap; gap:10px; margin-top:14px; }}
.pill {{ display:inline-block; background:rgba(255,255,255,.14); border:1px solid rgba(255,255,255,.26); border-radius:999px; padding:6px 10px; font-size:13px; }}
.cards {{ display:grid; grid-template-columns:repeat(4,1fr); gap:14px; margin:18px 0; }}
.card {{ background:white; border:1px solid var(--line); border-radius:8px; padding:16px; }}
.card .label {{ color:#667085; font-size:13px; }}
.card .value {{ margin-top:8px; font-size:24px; font-weight:700; }}
.grid {{ display:grid; grid-template-columns:1fr 1fr; gap:16px; }}
.panel {{ background:white; border:1px solid var(--line); border-radius:8px; padding:18px; margin-bottom:16px; }}
.panel h2 {{ margin:0 0 14px; font-size:18px; color:#1f4e78; }}
.conclusion-list,.risk-list {{ margin:0; padding-left:20px; line-height:1.8; }}
.share-box {{ display:grid; grid-template-columns:1fr auto; gap:12px; align-items:start; }}
.share-text {{ width:100%; min-height:118px; resize:vertical; border:1px solid var(--line); border-radius:8px; padding:12px; font:13px/1.6 inherit; box-sizing:border-box; background:#f8fafc; }}
.copy-btn {{ border:0; border-radius:8px; background:#1f4e78; color:white; padding:11px 16px; font-weight:700; cursor:pointer; }}
.copy-btn:active {{ transform:translateY(1px); }}
.copy-status {{ color:#16794c; font-size:13px; margin-top:8px; min-height:18px; }}
.bar-row {{ display:grid; grid-template-columns:160px 1fr 62px; align-items:center; gap:10px; margin:7px 0; font-size:13px; }}
.bar-label {{ overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }}
.bar-track {{ height:14px; background:#eef2f6; border-radius:999px; overflow:hidden; }}
.bar-fill {{ display:block; height:100%; border-radius:999px; }}
table {{ width:100%; border-collapse:collapse; font-size:13px; }}
.table-scroll {{ overflow:auto; border-radius:8px; }}
th {{ background:#1f4e78; color:white; text-align:left; padding:10px; position:sticky; top:0; }}
td {{ border-bottom:1px solid var(--line); padding:9px 10px; vertical-align:middle; }}
tr:nth-child(even) td {{ background:#fafafa; }}
.up {{ color:var(--red); font-weight:700; }}
.down {{ color:var(--green); font-weight:700; }}
.stack {{ display:flex; height:12px; width:150px; border-radius:999px; overflow:hidden; background:#eef2f6; }}
.stack span {{ display:block; height:100%; }}
.scatter {{ width:100%; height:330px; }}
.scatter line {{ stroke:#98a2b3; stroke-width:1; }}
.scatter circle {{ fill:#2f80ed; fill-opacity:.68; stroke:white; stroke-width:1.5; }}
.scatter text {{ fill:#667085; font-size:13px; }}
.empty {{ color:#667085; background:#f8fafc; border:1px dashed #cbd5e1; padding:18px; border-radius:8px; }}
.note {{ color:#667085; font-size:13px; line-height:1.7; }}
@media (max-width: 860px) {{ .cards,.grid,.share-box {{ grid-template-columns:1fr; }} .wrap {{ padding:16px; }} .hero {{ padding:22px 18px; }} .card .value {{ font-size:20px; }} }}
</style>
</head>
<body>
<div class="wrap">
  <section class="hero">
    <h1>K线选股日报</h1>
    <p>{date_label} | 筛选模式：{html.escape(str(getattr(cfg, "SCREEN_MODE", "normal")))} | Top {html.escape(str(getattr(cfg, "TOP_N", 20)))}</p>
    <div class="meta">
      <span class="pill">生成时间：{html.escape(generated_at)}</span>
      <span class="pill">信号口径：收盘后数据</span>
      <span class="pill">交易假设：T+1 开盘观察</span>
    </div>
  </section>
  <section class="disclaimer">仅为量化规则跟踪和学习交流，不构成投资建议；历史表现不代表未来收益。</section>
  <section class="cards">
    <div class="card"><div class="label">入选数量</div><div class="value">{stats['pick_count']}只</div></div>
    <div class="card"><div class="label">最高分股票</div><div class="value">{html.escape(str(stats['top_name']))}</div></div>
    <div class="card"><div class="label">平均分</div><div class="value">{stats['avg_score']:.1f}</div></div>
    <div class="card"><div class="label">风险提示</div><div class="value">{stats['risk_count']}只</div></div>
  </section>
  <section class="grid">
    <div class="panel"><h2>今日结论</h2>{conclusion_html}</div>
    <div class="panel"><h2>风险提示</h2>{risk_html}</div>
  </section>
  <section class="panel">
    <h2>微信/聊天摘要</h2>
    <div class="share-box">
      <textarea id="shareText" class="share-text" readonly>{html.escape(share_text)}</textarea>
      <div>
        <button class="copy-btn" onclick="copyShareText()">复制摘要</button>
        <div id="copyStatus" class="copy-status"></div>
      </div>
    </div>
  </section>
  <section class="grid">
    <div class="panel"><h2>Top20 综合评分</h2>{_bar_svg(top_items)}</div>
    <div class="panel"><h2>ADX 与量比散点</h2>{_scatter_html(picks)}</div>
  </section>
  <section class="panel"><h2>技术筛选漏斗</h2>{_bar_svg(funnel_items, value_key='value', label_key='label') if funnel_items else "<div class='empty'>暂无技术诊断数据</div>"}</section>
  <section class="panel"><h2>候选股详情</h2><div class="table-scroll"><table><thead><tr><th>排名</th><th>代码</th><th>名称</th><th>现价</th><th>行业</th><th>总分</th><th>20日线上</th><th>MA20斜率</th><th>涨跌幅</th><th>ADX</th><th>量比</th><th>评分拆解</th><th>入场时机</th><th>退出计划</th><th>第1日计划</th><th>第2日计划</th><th>第3日计划</th></tr></thead><tbody>{''.join(table_rows)}</tbody></table></div></section>
  <section class="panel"><h2>准确性评估</h2>{accuracy_html}</section>
  <section class="panel note"><strong>说明：</strong>本报告仅用于规则选股跟踪，不构成投资建议。信号基于收盘后数据，实盘需考虑涨跌停、停牌、滑点、手续费和成交量约束。</section>
</div>
<script>
function copyShareText() {{
  const text = document.getElementById('shareText');
  const status = document.getElementById('copyStatus');
  text.focus();
  text.select();
  const value = text.value;
  if (navigator.clipboard && window.isSecureContext) {{
    navigator.clipboard.writeText(value).then(() => {{
      status.textContent = '已复制，可以直接粘贴发送。';
    }}).catch(() => fallbackCopy(text, status));
  }} else {{
    fallbackCopy(text, status);
  }}
}}
function fallbackCopy(text, status) {{
  try {{
    document.execCommand('copy');
    status.textContent = '已复制，可以直接粘贴发送。';
  }} catch (e) {{
    status.textContent = '如果浏览器拦截复制，请手动选中文本复制。';
  }}
}}
</script>
</body>
</html>"""
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(doc)
    print(f"[报告] HTML报告: {html_path}")
    return html_path


def _find_chrome_executable() -> Optional[str]:
    candidates = [
        "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
        "/Applications/Microsoft Edge.app/Contents/MacOS/Microsoft Edge",
        "/Applications/Chromium.app/Contents/MacOS/Chromium",
        "/Applications/Brave Browser.app/Contents/MacOS/Brave Browser",
        shutil.which("google-chrome"),
        shutil.which("chromium"),
        shutil.which("chromium-browser"),
        shutil.which("chrome"),
    ]
    for path in candidates:
        if path and os.path.exists(path):
            return path
    return None


def convert_html_to_pdf(html_path: str, pdf_path: str = None) -> Optional[str]:
    """把静态 HTML 报告打印为 PDF。失败时返回 None，不影响主流程。"""
    if not html_path or not os.path.exists(html_path):
        return None
    if pdf_path is None:
        pdf_path = os.path.splitext(html_path)[0] + ".pdf"

    chrome = _find_chrome_executable()
    if not chrome:
        print("[报告] 未找到 Chrome/Edge/Chromium，跳过 PDF。")
        return None

    html_uri = "file://" + os.path.abspath(html_path)

    try:
        from playwright.sync_api import sync_playwright

        with sync_playwright() as p:
            browser = p.chromium.launch(executable_path=chrome, headless=True)
            page = browser.new_page(viewport={"width": 1360, "height": 1800})
            page.goto(html_uri, wait_until="networkidle")
            page.emulate_media(media="screen")
            page.pdf(
                path=pdf_path,
                format="A4",
                print_background=True,
                margin={"top": "12mm", "right": "10mm", "bottom": "12mm", "left": "10mm"},
            )
            browser.close()
    except Exception:
        # 没有 Playwright 或浏览器协议失败时，退回 Chrome 原生命令。
        try:
            with tempfile.TemporaryDirectory(prefix="stock_pick_pdf_") as tmpdir:
                cmd = [
                    chrome,
                    "--headless=new",
                    "--disable-gpu",
                    "--no-first-run",
                    "--no-default-browser-check",
                    f"--user-data-dir={tmpdir}",
                    f"--print-to-pdf={pdf_path}",
                    "--print-to-pdf-no-header",
                    html_uri,
                ]
                subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=60)
        except Exception as e:
            print(f"[报告] PDF生成失败，已跳过: {e}")
            return None

    if os.path.exists(pdf_path) and os.path.getsize(pdf_path) > 0:
        print(f"[报告] PDF报告: {pdf_path}")
        return pdf_path
    print("[报告] PDF生成失败，输出文件为空。")
    return None


def generate_pick_reports(
    pick_df: pd.DataFrame = None,
    trade_date=None,
    picks_path: str = None,
    output_dir: str = None,
) -> dict:
    if trade_date is None and picks_path:
        m = re.search(r"stock_pick_(\d{8})", os.path.basename(picks_path))
        if m:
            trade_date = datetime.datetime.strptime(m.group(1), "%Y%m%d").date()
    output_dir = output_dir or output_day_dir(trade_date)
    if pick_df is None:
        if picks_path is None:
            picks_path = os.path.join(output_dir, f"stock_pick_{_date_key(trade_date)}.csv")
        if not os.path.exists(picks_path):
            raise FileNotFoundError(f"未找到选股结果文件: {picks_path}")
        pick_df = pd.read_csv(picks_path)
    pick_df = pick_df.copy()
    if "rank" not in pick_df.columns:
        pick_df.insert(0, "rank", range(1, len(pick_df) + 1))

    excel_path = generate_excel_report(pick_df, trade_date=trade_date, output_dir=output_dir)
    html_path = generate_html_report(pick_df, trade_date=trade_date, output_dir=output_dir)
    pdf_path = convert_html_to_pdf(html_path)
    manifest = {
        "created_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "trade_date": _date_label(trade_date),
        "excel": excel_path,
        "html": html_path,
        "pdf": pdf_path,
    }
    manifest_path = os.path.join(output_dir, f"stock_pick_report_{_date_key(trade_date)}.json")
    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)
    return manifest
